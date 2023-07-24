import pandas


# create a function to fetch data from xlsx file using the pandas module
def import_excel_data():
    # retrieve file from path to edit
    excel_data = pandas.read_excel('marks.xlsx',
                                   sheet_name='Sheet',)
    try:

        #  create a searching variable named search_surname
        search_surname = str(input('Search surname: '))
        surname_search = search_surname

        # create a variable to query data using surname_search
        query_var = excel_data.query("Surname == @surname_search")
        print(query_var)

        # find the index value from data frame
        row_index = next(iter(query_var[query_var['Surname'] == surname_search].index))
        print(row_index)

        # Create a function to edit a desired column of row
        def edit_row():
            import openpyxl
            wb = openpyxl.load_workbook('marks.xlsx')
            sheet = wb["Sheet"]
            # create a variable to use for retrieving cell values by adding row_index & with 2
            cell_value = row_index + 2
            # Manually get each cell value name using the row index
            row_name = sheet[f'A{cell_value}'].value
            row_surname = sheet[f'B{cell_value}'].value
            pr_marks = sheet[f'C{cell_value}'].value
            us_marks = sheet[f'D{cell_value}'].value
            ur_marks = sheet[f'E{cell_value}'].value

            # set new values to desired cells from user input
            sheet[f'A{cell_value}'].value = input('Edit name: ')
            sheet[f'B{cell_value}'].value = input('Edit surname: ')
            sheet[f'C{cell_value}'].value = input('Edit prepared reading marks: ')
            sheet[f'D{cell_value}'].value = input('Edit unprepared speech marks: ')
            sheet[f'E{cell_value}'].value = input('Edit unprepared reading marks: ')

            # new_data = {f'{sheet[f"A{cell_value}"].value : str(input("Edit name: "))}'}

            print('The following row: ', row_name, row_surname, pr_marks, us_marks, ur_marks, '\n')
            print('has been successfully updated')

            wb.save('marks.xlsx')

            # sheet.cell(row=1, column=1).value = 'Pimp'
            # sheet.cell(row=1, column=2).value = "Dealer"

        edit_row()
        # df = pd.DataFrame(colors)
        # df['first_set'] = df['Name'].replace(['Blue'], 'Green')

        return query_var
    except StopIteration as e:
        print(e)
        print("Handled stopIteration \n No name edited. ")


if __name__ == "__main__":

    import_excel_data()
